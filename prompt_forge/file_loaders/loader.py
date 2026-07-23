"""
File loader abstraction.

Ships with sensible defaults for common file types. Users can register
custom loaders for any extension.
"""

from __future__ import annotations

import json
import csv
import dataclasses
from pathlib import Path
from typing import Callable


@dataclasses.dataclass
class FileContent:
    """Loaded file content, ready to be injected into LLM context."""

    text: str
    source_path: str
    file_type: str
    metadata: dict | None = None


# Type alias for a loader function: takes a Path, returns text content.
LoaderFn = Callable[[Path], str]


# ── Built-in loaders ──────────────────────────────────────────────────

def _load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _load_json(path: Path) -> str:
    data = json.loads(path.read_text(encoding="utf-8"))
    return json.dumps(data, indent=2, ensure_ascii=False)


def _load_csv(path: Path) -> str:
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return ""
    # Render as a readable table
    lines = []
    for row in rows:
        lines.append(" | ".join(row))
    return "\n".join(lines)


def _load_pdf(path: Path) -> str:
    """PDF loader: pdfplumber with pytesseract OCR fallback."""
    try:
        import pdfplumber
    except ImportError:
        raise ImportError(
            "pdfplumber is required for PDF loading. "
            "Install it with: pip install pdfplumber"
        )

    texts: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text.strip():
                texts.append(page_text)
            else:
                # OCR fallback
                texts.append(_ocr_page(page))

    return "\n\n".join(texts)


def _ocr_page(page) -> str:
    """OCR a single pdfplumber page using pytesseract."""
    try:
        import pytesseract
    except ImportError:
        return "[OCR unavailable — install pytesseract and Pillow]"

    img = page.to_image(resolution=300).original
    return pytesseract.image_to_string(img)


def _load_excel(path: Path) -> str:
    """Load .xlsx/.xls as text tables."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel loading. "
            "Install it with: pip install openpyxl"
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(" | ".join(str(c) if c is not None else "" for c in row))
        parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def _load_docx(path: Path) -> str:
    """Load .docx as plain text."""
    try:
        import docx
    except ImportError:
        raise ImportError(
            "python-docx is required for .docx loading. "
            "Install it with: pip install python-docx"
        )
    doc = docx.Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def _load_image(path: Path) -> str:
    """OCR an image file using pytesseract."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return f"[Image file: {path.name} — install pytesseract + Pillow for OCR]"

    img = Image.open(path)
    return pytesseract.image_to_string(img)


# ── Default extension → loader mapping ────────────────────────────────

_DEFAULT_LOADERS: dict[str, LoaderFn] = {
    ".txt": _load_text,
    ".md": _load_text,
    ".log": _load_text,
    ".xml": _load_text,
    ".html": _load_text,
    ".htm": _load_text,
    ".yaml": _load_text,
    ".yml": _load_text,
    ".toml": _load_text,
    ".ini": _load_text,
    ".cfg": _load_text,
    ".py": _load_text,
    ".js": _load_text,
    ".ts": _load_text,
    ".json": _load_json,
    ".jsonl": _load_text,
    ".csv": _load_csv,
    ".tsv": _load_csv,
    ".pdf": _load_pdf,
    ".xlsx": _load_excel,
    ".xls": _load_excel,
    ".docx": _load_docx,
    ".png": _load_image,
    ".jpg": _load_image,
    ".jpeg": _load_image,
    ".tiff": _load_image,
    ".bmp": _load_image,
}


# ── FileLoader class ─────────────────────────────────────────────────

class FileLoader:
    """
    Loads files into text content suitable for LLM context.

    Uses built-in loaders for common extensions. Users can register
    custom loaders for any extension.

    Usage:
        loader = FileLoader()
        loader.register(".parquet", my_parquet_loader)
        content = loader.load("path/to/file.pdf")
    """

    def __init__(self):
        self._loaders: dict[str, LoaderFn] = dict(_DEFAULT_LOADERS)

    def register(self, extension: str, loader_fn: LoaderFn) -> None:
        """
        Register a custom loader for a file extension.

        Args:
            extension: File extension including dot (e.g., ".parquet")
            loader_fn: Callable that takes a Path and returns str.
        """
        if not extension.startswith("."):
            extension = f".{extension}"
        self._loaders[extension] = loader_fn

    def can_load(self, path: str | Path) -> bool:
        """Check if a loader exists for the given file type."""
        ext = Path(path).suffix.lower()
        return ext in self._loaders

    def load(self, path: str | Path) -> FileContent:
        """
        Load a file and return its text content.

        Args:
            path: Path to the file.

        Returns:
            FileContent with the extracted text.

        Raises:
            ValueError: If no loader is registered for the file extension.
        """
        path = Path(path)
        ext = path.suffix.lower()
        loader_fn = self._loaders.get(ext)
        if loader_fn is None:
            raise ValueError(
                f"No loader registered for '{ext}'. "
                f"Register one with loader.register('{ext}', my_loader_fn). "
                f"Supported: {sorted(self._loaders.keys())}"
            )
        text = loader_fn(path)
        return FileContent(
            text=text,
            source_path=str(path),
            file_type=ext,
            metadata={"size_bytes": path.stat().st_size},
        )

    @property
    def supported_extensions(self) -> list[str]:
        """List all supported file extensions."""
        return sorted(self._loaders.keys())


def get_default_loader() -> FileLoader:
    """Get a FileLoader pre-configured with all built-in loaders."""
    return FileLoader()
